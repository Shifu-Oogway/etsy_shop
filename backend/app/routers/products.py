from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.product import Product
from app.schemas.product import ProductCreate, ProductOut, ProductUpdate

router = APIRouter(prefix="/products", tags=["products"])


@router.get("", response_model=list[ProductOut])
async def list_products(limit: int = 50, offset: int = 0, db: AsyncSession = Depends(get_db)):
    rows = await db.execute(select(Product).order_by(Product.id.desc()).limit(limit).offset(offset))
    return rows.scalars().all()


@router.get("/count")
async def count_products(db: AsyncSession = Depends(get_db)):
    total = (await db.execute(select(func.count(Product.id)))).scalar() or 0
    return {"total": total}


@router.get("/{product_id}", response_model=ProductOut)
async def get_product(product_id: int, db: AsyncSession = Depends(get_db)):
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(404, f"product {product_id} not found")
    return product


@router.post("", response_model=ProductOut, status_code=201)
async def create_product(body: ProductCreate, db: AsyncSession = Depends(get_db)):
    product = Product(**body.model_dump())
    db.add(product)
    await db.commit()
    await db.refresh(product)
    return product


@router.patch("/{product_id}", response_model=ProductOut)
async def update_product(product_id: int, body: ProductUpdate, db: AsyncSession = Depends(get_db)):
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(404, f"product {product_id} not found")
    for key, value in body.model_dump(exclude_none=True).items():
        setattr(product, key, value)
    await db.commit()
    await db.refresh(product)
    return product


@router.delete("/{product_id}", status_code=204)
async def delete_product(product_id: int, db: AsyncSession = Depends(get_db)):
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(404, f"product {product_id} not found")
    await db.delete(product)
    await db.commit()


import mimetypes
import os
from fastapi.responses import FileResponse, HTMLResponse


@router.get("/{product_id}/download")
async def download_product(product_id: int, db: AsyncSession = Depends(get_db)):
    """Download the generated product file directly."""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(404, f"product {product_id} not found")
    if not product.file_path or not os.path.exists(product.file_path):
        raise HTTPException(404, "file not yet generated")
    mime, _ = mimetypes.guess_type(product.file_path)
    return FileResponse(
        path=product.file_path,
        media_type=mime or "application/octet-stream",
        filename=os.path.basename(product.file_path),
    )


@router.get("/{product_id}/preview", response_class=HTMLResponse)
async def preview_product(product_id: int, db: AsyncSession = Depends(get_db)):
    """Render an HTML preview of the product — embeddable in an iframe."""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(404, f"product {product_id} not found")

    file_path = product.file_path
    spec = product.spec or {}

    # ── PDF ──────────────────────────────────────────────────────────────────
    if file_path and file_path.endswith(".pdf") and os.path.exists(file_path):
        import base64
        data = base64.b64encode(open(file_path, "rb").read()).decode()
        return HTMLResponse(_pdf_preview(product.title, data))

    # ── Excel / CSV ──────────────────────────────────────────────────────────
    if file_path and (file_path.endswith(".xlsx") or file_path.endswith(".csv")):
        return HTMLResponse(_excel_preview(product, file_path))

    # ── Markdown (Notion) ────────────────────────────────────────────────────
    if file_path and file_path.endswith(".md") and os.path.exists(file_path):
        text = open(file_path, encoding="utf-8").read()
        return HTMLResponse(_markdown_preview(product.title, text))

    # ── Spec-only fallback (not yet generated) ───────────────────────────────
    return HTMLResponse(_spec_preview(product, spec))


# ── preview renderers ────────────────────────────────────────────────────────

_STYLE = """
  body { margin:0; font-family:'Segoe UI',system-ui,sans-serif; background:#0f1117; color:#e4e8f0; }
  .wrap { padding: 28px 32px; max-width: 860px; margin: 0 auto; }
  h1 { font-size:20px; font-weight:700; margin-bottom:6px; }
  h2 { font-size:14px; font-weight:600; color:#6b7a99; text-transform:uppercase;
       letter-spacing:.06em; margin:24px 0 10px; border-top:1px solid #252c3a; padding-top:16px; }
  h3 { font-size:14px; font-weight:600; margin:14px 0 6px; }
  p, li { font-size:14px; color:#b0bbd0; line-height:1.6; }
  ul { padding-left:18px; }
  .meta { font-size:12px; color:#6b7a99; margin-bottom:20px; }
  .badge { display:inline-block; padding:2px 8px; border-radius:999px; font-size:11px;
           font-weight:600; background:rgba(79,142,247,.15); color:#4f8ef7; }
  table { width:100%; border-collapse:collapse; font-size:13px; margin-top:8px; }
  th { text-align:left; padding:8px 10px; font-size:11px; font-weight:600; color:#6b7a99;
       text-transform:uppercase; letter-spacing:.06em; border-bottom:1px solid #252c3a; }
  td { padding:8px 10px; border-bottom:1px solid #1e2330; color:#b0bbd0; }
  .sheet-name { font-size:13px; font-weight:600; margin:18px 0 6px; color:#e4e8f0; }
  pre { background:#1e2330; border:1px solid #252c3a; border-radius:8px; padding:16px;
        font-size:13px; white-space:pre-wrap; color:#b0bbd0; }
  .page-block { background:#1e2330; border:1px solid #252c3a; border-radius:10px;
                padding:16px 20px; margin-bottom:12px; }
  .section-row { padding:6px 0; border-bottom:1px solid #252c3a; font-size:13px;
                 color:#b0bbd0; display:flex; align-items:center; gap:8px; }
  .section-row:last-child { border-bottom:none; }
  .bullet { color:#4f8ef7; font-size:16px; }
  .not-generated { text-align:center; padding:60px 20px; color:#6b7a99; }
  .not-generated .icon { font-size:40px; margin-bottom:12px; }
"""


def _pdf_preview(title: str, b64: str) -> str:
    return f"""<!doctype html><html><head><meta charset=utf-8>
<title>{title}</title><style>{_STYLE}
  embed {{ width:100%; height:calc(100vh - 56px); border:none; border-radius:8px; }}
  .pdf-wrap {{ padding:16px; }}
</style></head><body>
<div class="pdf-wrap">
  <embed src="data:application/pdf;base64,{b64}" type="application/pdf" />
</div></body></html>"""


def _excel_preview(product, file_path: str) -> str:
    rows_html = ""
    try:
        if file_path.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for name in wb.sheetnames:
                ws = wb[name]
                rows_html += f'<div class="sheet-name">Sheet: {name}</div><table>'
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    tag = "th" if i == 0 else "td"
                    cells = "".join(f"<{tag}>{c if c is not None else ''}</{tag}>" for c in row)
                    rows_html += f"<tr>{cells}</tr>"
                    if i > 30:
                        rows_html += f"<tr><td colspan='{len(row)}' style='color:#6b7a99'>…truncated</td></tr>"
                        break
                rows_html += "</table>"
        else:  # CSV
            import csv
            with open(file_path, encoding="utf-8") as f:
                reader = csv.reader(f)
                rows_html += "<table>"
                for i, row in enumerate(reader):
                    tag = "th" if i == 0 else "td"
                    cells = "".join(f"<{tag}>{c}</{tag}>" for c in row)
                    rows_html += f"<tr>{cells}</tr>"
                    if i > 30:
                        break
                rows_html += "</table>"
    except Exception as exc:
        rows_html = f"<pre>Could not parse file: {exc}</pre>"

    return f"""<!doctype html><html><head><meta charset=utf-8>
<title>{product.title}</title><style>{_STYLE}</style></head><body>
<div class="wrap">
  <h1>{product.title}</h1>
  <div class="meta"><span class="badge">{product.product_type}</span>
    &nbsp;·&nbsp;{product.niche}&nbsp;·&nbsp;${product.price:.2f}</div>
  {rows_html}
</div></body></html>"""


def _markdown_preview(title: str, text: str) -> str:
    # Convert basic markdown to HTML without deps
    import re
    html = ""
    for line in text.splitlines():
        if line.startswith("### "): html += f"<h3>{line[4:]}</h3>\n"
        elif line.startswith("## "): html += f"<h2>{line[3:]}</h2>\n"
        elif line.startswith("# "):  html += f"<h1>{line[2:]}</h1>\n"
        elif line.startswith("- "):  html += f"<li>{line[2:]}</li>\n"
        elif line.strip() == "":    html += "<p></p>\n"
        else:                        html += f"<p>{line}</p>\n"
    return f"""<!doctype html><html><head><meta charset=utf-8>
<title>{title}</title><style>{_STYLE}</style></head><body>
<div class="wrap">{html}</div></body></html>"""


def _spec_preview(product, spec: dict) -> str:
    """Renders the product spec as a structured preview when no file exists yet."""
    ptype = product.product_type

    if ptype == "pdf_planner":
        pages = spec.get("pages", [])
        body = "".join(
            f"""<div class="page-block">
              <h3>📄 {p.get('title','Page')}</h3>
              {''.join(f'<div class="section-row"><span class="bullet">•</span>{s}</div>'
                       for s in p.get('sections',[]))}
            </div>""" for p in pages
        ) or '<div class="not-generated"><div class="icon">📄</div><p>No spec data yet</p></div>'

    elif ptype == "excel_template":
        sheets = spec.get("sheets", [])
        body = "".join(
            f"""<div class="sheet-name">Sheet: {s.get('name','Sheet')}</div>
            <table><tr>{''.join(f'<th>{h}</th>' for h in s.get('headers',[]))}</tr>
            <tr>{''.join(f'<td style="color:#6b7a99;font-style:italic">sample data</td>' for _ in s.get('headers',[]))}</tr>
            </table>""" for s in sheets
        ) or '<div class="not-generated"><div class="icon">📊</div><p>No spec data yet</p></div>'

    elif ptype == "notion_template":
        blocks = spec.get("blocks", [])
        body = "".join(
            f"""<div class="page-block">
              <h3>{b.get('heading','')}</h3>
              <p>{b.get('content','')}</p>
            </div>""" for b in blocks
        ) or '<div class="not-generated"><div class="icon">📝</div><p>No spec data yet</p></div>'

    else:
        body = f'<pre>{__import__("json").dumps(spec, indent=2)}</pre>'

    status_color = {"qa_passed":"#3ecf8e","published":"#3ecf8e","qa_failed":"#f2544a"}.get(product.status, "#f5a623")
    return f"""<!doctype html><html><head><meta charset=utf-8>
<title>{product.title}</title><style>{_STYLE}</style></head><body>
<div class="wrap">
  <h1>{product.title}</h1>
  <div class="meta">
    <span class="badge">{product.product_type}</span>
    &nbsp;·&nbsp;{product.niche or '—'}
    &nbsp;·&nbsp;${product.price:.2f}
    &nbsp;·&nbsp;<span style="color:{status_color}">{product.status}</span>
  </div>
  {body}
</div></body></html>"""


from pydantic import BaseModel as _BaseModel


class ProductPatch(_BaseModel):
    title: str | None = None
    description: str | None = None
    price: float | None = None
    niche: str | None = None


@router.patch("/{product_id}")
async def patch_product(product_id: int, body: ProductPatch,
                        db: AsyncSession = Depends(get_db)):
    """Edit a product before (re)publishing. Resets QA on content changes so
    edited products must pass the gate again."""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(404, f"product {product_id} not found")

    from app.models.product import ProductStatus
    content_changed = False
    if body.title is not None and body.title.strip():
        product.title = body.title.strip()[:255]
        content_changed = True
    if body.description is not None:
        product.description = body.description
        content_changed = True
    if body.price is not None:
        product.price = max(0.99, min(50.0, float(body.price)))
    if body.niche is not None:
        product.niche = body.niche[:120]

    if content_changed and product.status in (ProductStatus.qa_passed,):
        product.status = ProductStatus.generated  # must re-QA after edits

    await db.commit()
    return {"id": product.id, "title": product.title,
            "price": product.price, "status": product.status.value}


@router.get("/{product_id}/images/{index}")
async def product_image(product_id: int, index: int,
                        db: AsyncSession = Depends(get_db)):
    """Serve a generated mockup image."""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(404, f"product {product_id} not found")
    images = (product.spec or {}).get("images", [])
    if index < 0 or index >= len(images) or not os.path.exists(images[index]):
        raise HTTPException(404, "image not found")
    return FileResponse(images[index], media_type="image/png")
