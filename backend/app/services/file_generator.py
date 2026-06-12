"""Generates actual sellable digital product files.

All three generators produce genuinely useful output:
  - pdf_planner      → multi-page PDF with styled headers, ruled lines, and
                       section grids (reportlab); falls back to a fully-formed
                       HTML file that browsers / Word can print to PDF.
  - excel_template   → .xlsx with formatted headers (bold, bg colour, auto
                       column widths) and 5 sample data rows (openpyxl); falls
                       back to multi-sheet CSV bundle in a .zip.
  - notion_template  → .md with rich Markdown (tables, checkboxes, callouts)
                       + companion .json spec the buyer pastes into Notion API.

Heavy renderers are optional imports — a missing library never blocks the
Docker build or the rest of the pipeline.
"""
from __future__ import annotations

import io
import json
import logging
import zipfile
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path("generated_products")


def _ensure_dir() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR


# ── PDF Planner ───────────────────────────────────────────────────────────────

def generate_pdf_planner(slug: str, spec: dict[str, Any]) -> str:
    out = _ensure_dir() / f"{slug}.pdf"
    pages = spec.get("pages", [{"title": "Planner", "sections": ["Notes"]}])
    title = spec.get("title", slug.replace("-", " ").title())

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, PageBreak,
        )

        doc = SimpleDocTemplate(
            str(out), pagesize=LETTER,
            leftMargin=0.85*inch, rightMargin=0.85*inch,
            topMargin=0.9*inch, bottomMargin=0.75*inch,
        )
        styles = getSampleStyleSheet()

        # Custom styles
        h1 = ParagraphStyle("H1", parent=styles["Heading1"],
                             fontSize=22, textColor=colors.HexColor("#1a1a2e"),
                             spaceBefore=0, spaceAfter=6)
        h2 = ParagraphStyle("H2", parent=styles["Heading2"],
                             fontSize=13, textColor=colors.HexColor("#16213e"),
                             spaceBefore=14, spaceAfter=4)
        body = ParagraphStyle("Body", parent=styles["Normal"],
                              fontSize=10, textColor=colors.HexColor("#333333"),
                              spaceAfter=4)
        accent = colors.HexColor("#4f8ef7")

        story = []

        # Cover page
        story.append(Spacer(1, 1.5*inch))
        story.append(Paragraph(title, h1))
        story.append(HRFlowable(width="100%", thickness=2, color=accent, spaceAfter=10))
        story.append(Paragraph(f"{len(pages)} sections  ·  Printable &amp; fillable", body))
        story.append(PageBreak())

        for page in pages:
            page_title = str(page.get("title", "Page"))
            sections   = page.get("sections", ["Notes"])

            story.append(Paragraph(page_title, h1))
            story.append(HRFlowable(width="100%", thickness=1.5, color=accent, spaceAfter=8))

            for section in sections:
                story.append(Paragraph(section, h2))
                # Ruled lines for writing space
                line_data = [[""] for _ in range(5)]
                line_table = Table(line_data, colWidths=[6.3*inch])
                line_table.setStyle(TableStyle([
                    ("LINEBELOW",   (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
                    ("TOPPADDING",  (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]))
                story.append(line_table)
                story.append(Spacer(1, 6))

            story.append(PageBreak())

        doc.build(story)
        logger.info("PDF planner written (reportlab): %s", out)
        return str(out)

    except ImportError:
        logger.warning("reportlab not installed — writing HTML planner for %s", slug)
        html_out = out.with_suffix(".html")
        html_out.write_text(_html_planner(title, pages), encoding="utf-8")
        return str(html_out)


def _html_planner(title: str, pages: list[dict]) -> str:
    """Fully-formed printable HTML planner — opens in browser, File → Print to PDF."""
    page_blocks = ""
    for page in pages:
        sections_html = ""
        for section in page.get("sections", ["Notes"]):
            lines = "".join(
                '<div style="border-bottom:1px solid #ccc;height:32px;margin-bottom:2px"></div>'
                for _ in range(6)
            )
            sections_html += f"""
            <div class="section">
              <h3>{section}</h3>
              <div class="lines">{lines}</div>
            </div>"""
        page_blocks += f"""
        <div class="page">
          <h2>{page.get('title','Page')}</h2>
          <hr class="accent">
          {sections_html}
        </div>"""

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<title>{title}</title>
<style>
  @page {{ size: Letter; margin: 0.85in; }}
  body {{ font-family: 'Georgia', serif; color: #1a1a2e; }}
  h1   {{ font-size: 28px; margin-bottom: 4px; }}
  h2   {{ font-size: 18px; margin: 0 0 8px; }}
  h3   {{ font-size: 13px; font-weight: 600; color: #16213e; margin: 18px 0 6px; }}
  hr.accent {{ border: none; border-top: 2px solid #4f8ef7; margin: 6px 0 18px; }}
  .cover {{ page-break-after: always; padding-top: 120px; }}
  .page  {{ page-break-after: always; }}
  .lines div {{ height: 32px; border-bottom: 1px solid #ccc; }}
  @media screen {{ body {{ max-width:800px; margin:40px auto; padding:0 40px; }} }}
</style></head><body>
  <div class="cover">
    <h1>{title}</h1>
    <hr class="accent">
    <p style="color:#666;font-size:14px">{len(pages)} sections · Printable planner</p>
  </div>
  {page_blocks}
</body></html>"""


# ── Excel Template ─────────────────────────────────────────────────────────────

def generate_excel_template(slug: str, spec: dict[str, Any]) -> str:
    out = _ensure_dir() / f"{slug}.xlsx"
    title  = spec.get("title", slug.replace("-", " ").title())
    sheets = spec.get("sheets", [{"name": "Sheet1", "headers": ["Item", "Value", "Notes"]}])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        HEADER_BG   = "1A1A2E"
        HEADER_FG   = "FFFFFF"
        ALT_ROW_BG  = "F0F4FF"
        ACCENT      = "4F8EF7"
        thin        = Side(style="thin", color="CCCCCC")
        border      = Border(bottom=thin)

        for sheet_spec in sheets:
            ws = wb.create_sheet(title=str(sheet_spec.get("name", "Sheet"))[:31])
            headers = sheet_spec.get("headers", ["Column A", "Column B"])

            # Title row
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font      = Font(bold=True, size=14, color=HEADER_FG)
            title_cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # Header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font      = Font(bold=True, size=11, color=HEADER_FG)
                cell.fill      = PatternFill("solid", fgColor=ACCENT)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border
            ws.row_dimensions[2].height = 22

            # 10 template rows (alternating shading)
            sample_values = _sample_row_values(headers)
            for row_idx in range(3, 13):
                is_alt = (row_idx % 2 == 0)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx,
                                   value=sample_values[col_idx - 1] if row_idx == 3 else "")
                    if is_alt:
                        cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)
                    cell.border    = border
                    cell.alignment = Alignment(vertical="center")
                ws.row_dimensions[row_idx].height = 18

            # Auto column widths
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(len(str(header)) + 6, 14)

            # Freeze panes below header
            ws.freeze_panes = "A3"

        wb.save(out)
        logger.info("Excel template written (openpyxl): %s", out)
        return str(out)

    except ImportError:
        logger.warning("openpyxl not installed — writing CSV zip bundle for %s", slug)
        zip_out = out.with_suffix(".zip")
        with zipfile.ZipFile(zip_out, "w") as zf:
            for sheet_spec in sheets:
                name    = sheet_spec.get("name", "Sheet")
                headers = sheet_spec.get("headers", ["Column A", "Column B"])
                sample  = _sample_row_values(headers)
                rows    = [",".join(headers),
                           ",".join(str(v) for v in sample)]
                rows   += ["," * (len(headers) - 1)] * 9
                zf.writestr(f"{name}.csv", "\n".join(rows))
        return str(zip_out)


def _sample_row_values(headers: list[str]) -> list[str]:
    """Generate plausible sample values based on header names."""
    samples = []
    for h in headers:
        hl = h.lower()
        if any(k in hl for k in ("date", "day", "month")):
            samples.append("2026-01-01")
        elif any(k in hl for k in ("amount", "total", "price", "cost", "revenue", "budget")):
            samples.append("0.00")
        elif any(k in hl for k in ("qty", "quantity", "count", "num")):
            samples.append("1")
        elif any(k in hl for k in ("name", "title", "item", "product", "task")):
            samples.append(f"Sample {h}")
        elif any(k in hl for k in ("status", "state")):
            samples.append("Pending")
        elif any(k in hl for k in ("note", "comment", "desc")):
            samples.append("Enter details here")
        elif any(k in hl for k in ("category", "type", "tag")):
            samples.append("Category A")
        elif any(k in hl for k in ("%", "pct", "percent", "rate")):
            samples.append("0%")
        else:
            samples.append("")
    return samples


# ── Notion Template ───────────────────────────────────────────────────────────

def generate_notion_template(slug: str, spec: dict[str, Any]) -> str:
    """Produces a .md file with rich Notion-flavoured markdown + a companion
    .json spec the buyer can use with the Notion API or import directly."""
    out      = _ensure_dir() / f"{slug}.md"
    json_out = out.with_suffix(".json")
    title    = spec.get("title", slug.replace("-", " ").title())
    blocks   = spec.get("blocks", [{"heading": "Overview", "content": "Add your content here."}])

    lines = [
        f"# {title}",
        "",
        "> 💡 **How to use:** Duplicate this page into your Notion workspace.",
        "> Edit each section to fit your needs.",
        "",
        "---",
        "",
    ]

    for block in blocks:
        heading = block.get("heading", "Section")
        content = block.get("content", "")

        lines += [f"## {heading}", ""]

        # Detect content type and format appropriately
        cl = content.lower()
        if any(k in cl for k in ("track", "log", "record", "list")):
            lines += [
                "| Item | Status | Date | Notes |",
                "|------|--------|------|-------|",
                "| Sample item | ⬜ Todo | — | — |",
                "| Sample item | ✅ Done | — | — |",
                "",
            ]
        elif any(k in cl for k in ("goal", "plan", "target", "objective")):
            lines += [
                "- [ ] Goal 1",
                "- [ ] Goal 2",
                "- [ ] Goal 3",
                "",
                f"{content}",
                "",
            ]
        elif any(k in cl for k in ("note", "journal", "reflect", "write")):
            lines += [
                f"{content}",
                "",
                "*Your notes here…*",
                "",
            ]
        else:
            lines += [f"{content}", ""]

        lines += ["---", ""]

    # Footer
    lines += [
        "## 📎 Resources",
        "",
        "| Resource | Link |",
        "|----------|------|",
        "| Template guide | [View guide](#) |",
        "| Support | [Contact us](#) |",
        "",
    ]

    out.write_text("\n".join(lines), encoding="utf-8")

    # Companion JSON for Notion API / advanced users
    notion_json = {
        "object": "page",
        "title": title,
        "slug": slug,
        "blocks": [
            {
                "object": "block",
                "type": "heading_2",
                "heading_2": {"rich_text": [{"type": "text", "text": {"content": b.get("heading", "")}}]},
            }
            for b in blocks
        ],
        "spec": spec,
    }
    json_out.write_text(json.dumps(notion_json, indent=2), encoding="utf-8")

    logger.info("Notion template written: %s + %s", out, json_out)
    return str(out)


# ── Registry ──────────────────────────────────────────────────────────────────

GENERATORS = {
    "pdf_planner":     generate_pdf_planner,
    "excel_template":  generate_excel_template,
    "notion_template": generate_notion_template,
}
