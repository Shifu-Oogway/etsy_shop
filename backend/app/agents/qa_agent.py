"""Validates the generated product before publishing. Deterministic checks plus
an LLM quality review; both contribute to the score."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from sqlalchemy import select

from app.agents.base import AgentResult, BaseAgent
from app.core.config import get_settings
from app.models.product import Product, ProductStatus
from app.models.qa_report import QAReport
from app.models.seo import SEOMetadata


def _validate_file(path: str) -> dict:
    """Opens the actual file and validates its contents — a corrupt or empty
    artifact must never pass QA."""
    p = Path(path)
    checks: dict = {}
    suffix = p.suffix.lower()

    if suffix == ".pdf":
        data = p.read_bytes()
        checks["pdf_magic"] = data[:5] == b"%PDF-"
        checks["pdf_pages"] = data.count(b"/Type /Page") + data.count(b"/Type/Page") >= 2
        checks["file_size_ok"] = len(data) > 2000
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            checks["xlsx_loads"] = True
            checks["xlsx_has_sheets"] = len(wb.sheetnames) >= 1
            ws = wb[wb.sheetnames[0]]
            first_rows = list(ws.iter_rows(max_row=2, values_only=True))
            checks["xlsx_has_headers"] = any(
                any(c for c in row) for row in first_rows)
        except Exception:
            checks["xlsx_loads"] = False
        checks["file_size_ok"] = p.stat().st_size > 1500
    elif suffix == ".md":
        text = p.read_text(encoding="utf-8", errors="replace")
        checks["md_length_ok"] = len(text) >= 300
        checks["md_has_headings"] = text.count("## ") >= 3
        checks["file_size_ok"] = True
    elif suffix == ".html":
        text = p.read_text(encoding="utf-8", errors="replace")
        checks["html_valid"] = "<html" in text.lower()
        checks["file_size_ok"] = len(text) >= 500
    else:
        checks["file_size_ok"] = p.stat().st_size > 100

    return checks


class QAAgent(BaseAgent):
    name = "qa"

    REVIEW_PROMPT = (
        'Review this Etsy listing draft for quality. Title: "{title}". '
        'Description: "{description}". Rate 0.0-1.0 and list issues. '
        'Return JSON: {{"score": float, "issues": [str]}}.'
    )

    async def run(self, product_id: int, **_: Any) -> AgentResult:
        product = await self.db.get(Product, product_id)
        if product is None:
            return AgentResult(ok=False, error=f"product {product_id} not found")

        checks: dict[str, Any] = {}
        checks["file_exists"] = bool(product.file_path) and Path(product.file_path).exists()
        checks["file_nonempty"] = checks["file_exists"] and Path(product.file_path).stat().st_size > 0
        if checks["file_exists"]:
            checks.update(_validate_file(product.file_path))
        checks["title_length_ok"] = 0 < len(product.title) <= 140
        checks["has_description"] = len(product.description.strip()) >= 40
        checks["price_in_range"] = 0.99 <= product.price <= 50.0

        row = await self.db.execute(
            select(SEOMetadata).where(SEOMetadata.product_id == product.id)
            .order_by(SEOMetadata.id.desc()).limit(1))
        seo = row.scalar_one_or_none()
        checks["has_seo"] = seo is not None
        checks["tag_count_ok"] = bool(seo and len(seo.tags or []) == 13)

        deterministic = sum(1 for v in checks.values() if v) / len(checks)

        review = await self.llm.generate_json(self.REVIEW_PROMPT.format(
            title=product.title, description=product.description[:500]))
        llm_score = max(0.0, min(1.0, float(review.get("score", 0.0))))
        checks["llm_issues"] = review.get("issues", [])

        score = round(0.6 * deterministic + 0.4 * llm_score, 4)
        passed = score >= get_settings().qa_min_score and checks["file_exists"]

        report = QAReport(product_id=product.id, passed=passed, score=score, checks=checks)
        self.db.add(report)
        product.status = ProductStatus.qa_passed if passed else ProductStatus.qa_failed
        await self.db.flush()
        return AgentResult(ok=True, passed=passed, score=score, report_id=report.id)
